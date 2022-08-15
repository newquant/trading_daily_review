from datetime import datetime

from openpyxl import Workbook

from robot.css import normal_style, highlight_style
from robot.mailbot import do_send
from robot.testbu import *
from robot.utils import *


def generate_xlsx(file_name: str) -> None:
    normal = normal_style()
    highlight = highlight_style()

    # 涨停
    limit_up_list = get_limit_up_list()
    # 炸板
    limit_up_broken_list = get_limit_up_broken_list()
    # 指数相关信息
    index_json = fetch_index_json()
    # 市场温度
    market_temperature = fetch_market_temperature()
    # 涨跌个数统计
    rise_fall_count = fetch_rise_fall_count()
    # 涨跌停个数统计
    limit_up_limit_down_count = fetch_limit_up_limit_down_count()
    # 涨停破板比例
    limit_up_count_limit_up_broken_ration = fetch_limit_up_count_limit_up_broken_ration()

    if len(limit_up_list) < 1:
        print("开盘准备阶段，请9：25后再生成 ")
    else:

        workbook = Workbook()
        sheet = workbook.create_sheet()

        sheet.cell(5 - 4, 1).value = "序号"
        sheet.cell(5 - 4, 2).value = "股票名称"
        sheet.cell(5 - 4, 3).value = "理由逻辑"
        sheet.cell(5 - 4, 4).value = "换手率"
        sheet.cell(5 - 4, 5).value = "首次封板"
        sheet.cell(5 - 4, 6).value = "最后封板"
        sheet.cell(5 - 4, 7).value = "开板数"
        sheet.cell(5 - 4, 8).value = "连板数"
        sheet.cell(5 - 4, 9).value = "价格"
        sheet.cell(5 - 4, 10).value = "涨幅"
        sheet.cell(5 - 4, 11).value = "封单比"
        sheet.cell(5 - 4, 12).value = "流通市值"
        sheet.cell(5 - 4, 13).value = "总市值"
        for idx, limit_up in enumerate(limit_up_list):
            sheet.cell(idx + 6 - 4, 1).value = idx + 1
            sheet.cell(idx + 6 - 4, 2).value = limit_up["stock_chi_name"] + "(" + limit_up["symbol"] + ")"
            if limit_up["surge_reason"] == "" or limit_up["surge_reason"] is None:
                sheet.cell(idx + 6 - 4, 3).value = '其他'
            else:
                sheet.cell(idx + 6 - 4, 3).value = ','.join(
                    [plates['plate_name'] for plates in limit_up['surge_reason']['related_plates']])
            # 换手率
            sheet.cell(idx + 6 - 4, 4).value = handle_percent(limit_up["turnover_ratio"])
            # 首次封板
            sheet.cell(idx + 6 - 4, 5).value = handle_date_time(limit_up["first_limit_up"])
            # 最后封板
            sheet.cell(idx + 6 - 4, 6).value = handle_date_time(limit_up["last_limit_up"])
            # 开板数
            sheet.cell(idx + 6 - 4, 7).value = limit_up["break_limit_up_times"]
            # 连板数
            sheet.cell(idx + 6 - 4, 8).value = limit_up["limit_up_days"]
            # 价格
            sheet.cell(idx + 6 - 4, 9).value = limit_up["price"]
            # 涨幅
            sheet.cell(idx + 6 - 4, 10).value = handle_percent(limit_up["change_percent"])
            # 封单比
            sheet.cell(idx + 6 - 4, 11).value = handle_percent(limit_up["buy_lock_volume_ratio"])
            # 流通市值
            sheet.cell(idx + 6 - 4, 12).value = handle_e(limit_up["non_restricted_capital"])
            # 总市值
            sheet.cell(idx + 6 - 4, 13).value = handle_e(limit_up["total_capital"])
            if limit_up["surge_reason"] == "" or limit_up["surge_reason"] is None:
                sheet.cell(idx + 6 - 4, 14).value = "其他"
            else:
                sheet.cell(idx + 6 - 4, 14).value = limit_up["surge_reason"]["stock_reason"]

        workbook.close()
        # 保存文件
        workbook.save(file_name)


if __name__ == '__main__':
    generate_xlsx(file_name="./data/复盘数据 " + datetime.now().strftime('%Y-%m-%dT%H:%M:%S') + ".xlsx")
