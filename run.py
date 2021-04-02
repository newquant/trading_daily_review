from datetime import datetime

from openpyxl import Workbook

from robot.css import normal_style, highlight_style
from robot.testbu import *
from robot.utils import *


def generate_xlsx(file_name: str) -> None:
    normal = normal_style()
    highlight = highlight_style()

    # 涨停
    limit_up_list = get_limit_up_list()
    # 炸板
    limit_up_broken_list = get_limit_up_broken_list()
    # 指数相关信息
    index_json = fetch_index_json()
    # 市场温度
    market_temperature = fetch_market_temperature()
    # 涨跌个数统计
    rise_fall_count = fetch_rise_fall_count()
    # 涨跌停个数统计
    limit_up_limit_down_count = fetch_limit_up_limit_down_count()
    # 涨停破板比例
    limit_up_count_limit_up_broken_ration = fetch_limit_up_count_limit_up_broken_ration()

    if len(limit_up_list) < 1:
        print("开盘准备阶段，请9：25后再生成 ")
    else:

        workbook = Workbook()
        sheet = workbook.create_sheet()

        # # +++++++++++++++++++++++++++++++++++++++++++++++++++++++
        sheet.cell(1, 1).value = '复盘数据'
        sheet.cell(1, 1).style = highlight
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        # # +++++++++++++++++++++++++++++++++++++++++++++++++++++++

        sheet.cell(2, 1).value = '日期'
        sheet.cell(2, 2).value = "上证指数"
        sheet.cell(2, 3).value = "深证成指"
        sheet.cell(2, 4).value = "创业板指"
        sheet.cell(2, 5).value = "市场热度"
        sheet.cell(2, 6).value = "涨跌比"
        sheet.cell(2, 7).value = "涨跌停比(不含新股)"
        sheet.cell(2, 8).value = "炸板数(炸板率）"

        sheet.cell(3, 1).value = datetime.now().strftime('%Y-%m-%d')

        ss000001 = index_json["000001.SS"]
        sheet.cell(3, 2).value = '{0}/{1}/{2}%'.format(round(ss000001[1], 2),
                                                       round(ss000001[2], 2),
                                                       round(ss000001[3], 2))
        sz399001 = index_json["399001.SZ"]
        sheet.cell(3, 3).value = '{0}/{1}/{2}%'.format(round(sz399001[1], 2),
                                                       round(sz399001[2], 2),
                                                       round(sz399001[3], 2))
        sz399006 = index_json["399006.SZ"]
        sheet.cell(3, 4).value = '{0}/{1}/{2}%'.format(round(sz399006[1], 2),
                                                       round(sz399006[2], 2),
                                                       round(sz399006[3], 2))

        if market_temperature is None:
            print("开盘准备阶段，请9：25后再生成")
        else:
            sheet.cell(3, 5).value = market_temperature

        sheet.cell(3, 6).value = rise_fall_count
        sheet.cell(3, 7).value = limit_up_limit_down_count
        sheet.cell(3, 8).value = limit_up_count_limit_up_broken_ration

        # # +++++++++++++++++++++++++++++++++++++++++++++++++
        sheet.cell(4, 1).value = "涨停池"
        sheet.cell(4, 1).style = highlight
        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=14)
        # # +++++++++++++++++++++++++++++++++++++++++++++++++

        sheet.cell(5, 1).value = "序号"
        sheet.cell(5, 2).value = "股票名称"
        sheet.cell(5, 3).value = "理由逻辑"
        sheet.cell(5, 4).value = "换手率"
        sheet.cell(5, 5).value = "首次封板"
        sheet.cell(5, 6).value = "最后封板"
        sheet.cell(5, 7).value = "开板数"
        sheet.cell(5, 8).value = "连板数"
        sheet.cell(5, 9).value = "价格"
        sheet.cell(5, 10).value = "涨幅"
        sheet.cell(5, 11).value = "封单比"
        sheet.cell(5, 12).value = "流通市值"
        sheet.cell(5, 13).value = "总市值"
        for idx, limit_up in enumerate(limit_up_list):
            sheet.cell(idx + 6, 1).value = idx + 1
            sheet.cell(idx + 6, 2).value = limit_up["stock_chi_name"] + "(" + limit_up["symbol"] + ")"
            if limit_up["surge_reason"] == "" or limit_up["surge_reason"] is None:
                sheet.cell(idx + 6, 3).value = '其他'
            else:
                sheet.cell(idx + 6, 3).value = ','.join(
                    [plates['plate_name'] for plates in limit_up['surge_reason']['related_plates']])
            # 换手率
            sheet.cell(idx + 6, 4).value = handle_percent(limit_up["turnover_ratio"])
            # 首次封板
            sheet.cell(idx + 6, 5).value = handle_date_time(limit_up["first_limit_up"])
            # 最后封板
            sheet.cell(idx + 6, 6).value = handle_date_time(limit_up["last_limit_up"])
            # 开板数
            sheet.cell(idx + 6, 7).value = limit_up["break_limit_up_times"]
            # 连板数
            sheet.cell(idx + 6, 8).value = limit_up["limit_up_days"]
            # 价格
            sheet.cell(idx + 6, 9).value = limit_up["price"]
            # 涨幅
            sheet.cell(idx + 6, 10).value = handle_percent(limit_up["change_percent"])
            # 封单比
            sheet.cell(idx + 6, 11).value = handle_percent(limit_up["buy_lock_volume_ratio"])
            # 流通市值
            sheet.cell(idx + 6, 12).value = handle_e(limit_up["non_restricted_capital"])
            # 总市值
            sheet.cell(idx + 6, 13).value = handle_e(limit_up["total_capital"])
            if limit_up["surge_reason"] == "" or limit_up["surge_reason"] is None:
                sheet.cell(idx + 6, 14).value = "其他"
            else:
                sheet.cell(idx + 6, 14).value = limit_up["surge_reason"]["stock_reason"]

        # # +++++++++++++++++++++++++++++++++++++++++++++++++++++++
        sheet.cell(6 + len(limit_up_list), 1).value = "炸板池"
        sheet.cell(6 + len(limit_up_list), 1).style = highlight
        sheet.merge_cells(start_row=6 + len(limit_up_list), start_column=1,
                          end_row=6 + len(limit_up_list), end_column=14)
        # # +++++++++++++++++++++++++++++++++++++++++++++++++++++++
        sheet.cell(7 + len(limit_up_list), 1).value = "序号"
        sheet.cell(7 + len(limit_up_list), 2).value = "股票名称"
        sheet.cell(7 + len(limit_up_list), 3).value = "理由逻辑"
        sheet.cell(7 + len(limit_up_list), 4).value = "换手率"
        sheet.cell(7 + len(limit_up_list), 5).value = "首次封板"
        sheet.cell(7 + len(limit_up_list), 6).value = "最后炸板"
        sheet.cell(7 + len(limit_up_list), 7).value = "开板数"
        sheet.cell(7 + len(limit_up_list), 8).value = "连板数"
        sheet.cell(7 + len(limit_up_list), 9).value = "价格"
        sheet.cell(7 + len(limit_up_list), 10).value = "涨跌幅"
        sheet.cell(7 + len(limit_up_list), 11).value = "流通市值"
        sheet.cell(7 + len(limit_up_list), 12).value = "总市值"
        if len(limit_up_broken_list) >= 1:
            for idx, limit_up_broken in enumerate(limit_up_broken_list):
                # 序号
                sheet.cell(8 + len(limit_up_list) + idx, 1).value = idx + 1
                # 股票名称
                sheet.cell(8 + len(limit_up_list) + idx, 2).value = limit_up_broken["stock_chi_name"] + "(" + \
                                                                    limit_up_broken["symbol"] + ")"
                # 理由逻辑
                if limit_up_broken["surge_reason"] == "":
                    sheet.cell(8 + len(limit_up_list) + idx, 3).value = "其他"
                else:
                    sheet.cell(8 + len(limit_up_list) + idx, 3).value = ','.join(
                        [plates['plate_name'] for plates in limit_up_broken['surge_reason']['related_plates']])
                # 换手率
                sheet.cell(8 + len(limit_up_list) + idx, 4).value = handle_percent(limit_up_broken["turnover_ratio"])
                # 首次封板
                sheet.cell(8 + len(limit_up_list) + idx, 5).value = handle_date_time(limit_up_broken["first_limit_up"])
                # 最后炸板
                sheet.cell(8 + len(limit_up_list) + idx, 6).value = handle_date_time(
                    limit_up_broken["last_break_limit_up"])
                # 开板数
                sheet.cell(8 + len(limit_up_list) + idx, 7).value = limit_up_broken["break_limit_up_times"]
                # 连板数
                sheet.cell(8 + len(limit_up_list) + idx, 8).value = limit_up_broken["limit_up_days"]
                # 价格
                sheet.cell(8 + len(limit_up_list) + idx, 9).value = limit_up_broken["price"]
                # 涨跌幅
                sheet.cell(8 + len(limit_up_list) + idx, 10).value = handle_percent(limit_up_broken["change_percent"])
                # 流通市值
                sheet.cell(8 + len(limit_up_list) + idx, 11).value = handle_e(limit_up_broken["non_restricted_capital"])
                # 总市值
                sheet.cell(8 + len(limit_up_list) + idx, 12).value = handle_e(limit_up_broken["total_capital"])

                if limit_up_broken["surge_reason"] == "":
                    sheet.cell(8 + len(limit_up_list) + idx, 13).value = "其他"
                else:
                    sheet.cell(8 + len(limit_up_list) + idx, 13).value = limit_up_broken["surge_reason"][
                        "stock_reason"]

        workbook.close()
        # 保存文件
        workbook.save(file_name)


if __name__ == '__main__':
    filename = 'data/复盘数据' + datetime.now().strftime('%Y-%m-%dT%H:%M:%S') + '.xlsx'
    generate_xlsx(file_name=filename)
